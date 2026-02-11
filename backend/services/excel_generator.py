"""
Excel Generator for Individual Support Plan (Litalico Format)

Generates Excel file based on assessment_v1 data structure
"""

from io import BytesIO
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter


def generate_support_plan_excel(
    session_data: dict,
    plan_id: str = None
) -> BytesIO:
    """
    Generate Individual Support Plan Excel file (2 sheets)

    Args:
        session_data: Dict containing assessment_result_v1 and child profile
        plan_id: Optional business_support_plans ID for user-edited data

    Returns:
        BytesIO: Excel file binary data

    Data Priority:
        1. If plan_id provided -> fetch from business_support_plans (2-column logic)
        2. If plan_id not provided -> fallback to assessment_v1 (backward compatibility)
    """
    wb = Workbook()

    # Sheet 1: Main Support Plan
    ws1 = wb.active
    ws1.title = "別紙1-1（個別支援計画書）"

    # Extract data
    assessment_v1 = extract_assessment_v1(session_data.get('assessment_result_v1', {}))

    if not assessment_v1:
        raise ValueError("assessment_v1 data not found")

    # Fetch plan data if plan_id provided
    plan_data = None
    if plan_id:
        from supabase import create_client
        import os

        try:
            supabase = create_client(
                os.getenv("SUPABASE_URL"),
                os.getenv("SUPABASE_ANON_KEY")
            )

            result = supabase.table('business_support_plans')\
                .select('*')\
                .eq('id', plan_id)\
                .single()\
                .execute()

            if result.data:
                plan_data = result.data
        except Exception as e:
            print(f"Warning: Failed to fetch plan data: {str(e)}")
            # Continue with assessment_v1 only

    # === Sheet 1: Main Support Plan ===
    generate_main_support_plan(ws1, assessment_v1, session_data, plan_data)

    # === Sheet 2: Support Details (Page 2) ===
    ws2 = wb.create_sheet(title="個別支援計画書（2/2ページ）")
    generate_support_details_page2(ws2, assessment_v1, session_data, plan_data)

    # === Sheet 3: Support Schedule (Appendix) ===
    ws3 = wb.create_sheet(title="別紙1-2（個別支援計画書別表）")
    generate_support_schedule(ws3, assessment_v1, session_data)

    # Save to BytesIO
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    return output


def generate_main_support_plan(
    ws,
    assessment_v1: dict,
    session_data: dict = None,
    plan_data: dict = None
):
    """
    Generate main support plan sheet

    Args:
        ws: Worksheet object
        assessment_v1: Assessment data (fallback)
        session_data: Session data (child profile, etc.)
        plan_data: Optional business_support_plans record (2-column structure)
    """

    # Set column widths
    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 50
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 25
    ws.column_dimensions['F'].width = 10

    # Styles
    header_font = Font(name='Meiryo UI', size=14, bold=True)
    normal_font = Font(name='Meiryo UI', size=10)
    small_font = Font(name='Meiryo UI', size=9)

    center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    left_align = Alignment(horizontal='left', vertical='top', wrap_text=True)

    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    header_fill = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')

    current_row = 1

    # Title
    ws.merge_cells(f'A{current_row}:F{current_row}')
    cell = ws[f'A{current_row}']
    cell.value = '個別支援計画書'
    cell.font = header_font
    cell.alignment = center_align
    current_row += 1

    # Child name and date - prioritize subjects table data
    child_name = session_data.get('subject_name', '〇〇 〇〇')
    child_age = session_data.get('subject_age', 5)

    # Fallback to assessment_v1 child_profile if subjects data not available
    if child_name == '〇〇 〇〇':
        child_profile = assessment_v1.get('child_profile', {})
        if child_profile.get('name'):
            child_name = child_profile.get('name')
        if child_profile.get('age'):
            child_age = child_profile.get('age')

    ws.merge_cells(f'A{current_row}:D{current_row}')
    ws[f'A{current_row}'] = f'利用児氏名：{child_name}（{child_age}歳）'
    ws[f'A{current_row}'].font = normal_font

    ws.merge_cells(f'E{current_row}:F{current_row}')
    ws[f'E{current_row}'] = f'作成年月日：{datetime.now().strftime("%Y年%m月%d日")}'
    ws[f'E{current_row}'].font = normal_font
    ws[f'E{current_row}'].alignment = Alignment(horizontal='right')
    current_row += 2

    # Family/Child Intentions
    if plan_data:
        child_intention = get_field_value(
            plan_data,
            'child_intention',
            assessment_v1.get('family_child_intentions', {}).get('child')
        )
        family_intention = get_field_value(
            plan_data,
            'family_intention',
            assessment_v1.get('family_child_intentions', {}).get('parents')
        )
    else:
        intentions = assessment_v1.get('family_child_intentions', {})
        child_intention = intentions.get('child')
        family_intention = intentions.get('parents')

    ws.merge_cells(f'A{current_row}:F{current_row}')
    cell = ws[f'A{current_row}']
    cell.value = '利用児及び家族の生活に対する意向'
    cell.font = Font(name='Meiryo UI', size=10, bold=True)
    cell.fill = header_fill
    cell.border = thin_border
    current_row += 1

    ws.merge_cells(f'A{current_row}:F{current_row}')
    cell = ws[f'A{current_row}']
    intention_text = []
    if child_intention:
        intention_text.append(f"• {child_intention}（本人）")
    if family_intention:
        intention_text.append(f"• {family_intention}（保護者）")
    cell.value = '\n'.join(intention_text) if intention_text else ''
    cell.font = normal_font
    cell.alignment = left_align
    cell.border = thin_border
    ws.row_dimensions[current_row].height = 40
    current_row += 2

    # Support Policy
    if plan_data:
        general_policy = get_field_value(
            plan_data,
            'general_policy',
            assessment_v1.get('support_policy', {}).get('child_understanding')
        )
    else:
        general_policy = assessment_v1.get('support_policy', {}).get('child_understanding', '')

    ws.merge_cells(f'A{current_row}:F{current_row}')
    cell = ws[f'A{current_row}']
    cell.value = '総合的な支援の方針'
    cell.font = Font(name='Meiryo UI', size=10, bold=True)
    cell.fill = header_fill
    cell.border = thin_border
    current_row += 1

    ws.merge_cells(f'A{current_row}:F{current_row}')
    cell = ws[f'A{current_row}']
    cell.value = general_policy if general_policy else ''
    cell.font = normal_font
    cell.alignment = left_align
    cell.border = thin_border
    ws.row_dimensions[current_row].height = 80
    current_row += 2

    # Long-term Goal
    if plan_data:
        lt_goal = get_field_value(
            plan_data,
            'long_term_goal',
            assessment_v1.get('long_term_goal', {}).get('goal')
        )
    else:
        lt_goal = assessment_v1.get('long_term_goal', {}).get('goal', '')

    ws.merge_cells(f'A{current_row}:F{current_row}')
    cell = ws[f'A{current_row}']
    cell.value = '長期目標（内容・期間等）'
    cell.font = Font(name='Meiryo UI', size=10, bold=True)
    cell.fill = header_fill
    cell.border = thin_border
    current_row += 1

    ws.merge_cells(f'A{current_row}:F{current_row}')
    cell = ws[f'A{current_row}']
    cell.value = lt_goal if lt_goal else ''
    cell.font = normal_font
    cell.alignment = left_align
    cell.border = thin_border
    ws.row_dimensions[current_row].height = 40
    current_row += 2

    # Short-term Goals
    if plan_data:
        short_term_goals = get_field_value(
            plan_data,
            'short_term_goals',
            assessment_v1.get('short_term_goals', [])
        )
    else:
        short_term_goals = assessment_v1.get('short_term_goals', [])

    # JSONB array safety check
    if not isinstance(short_term_goals, list):
        short_term_goals = []

    ws.merge_cells(f'A{current_row}:F{current_row}')
    cell = ws[f'A{current_row}']
    cell.value = '短期目標（内容・期間等）'
    cell.font = Font(name='Meiryo UI', size=10, bold=True)
    cell.fill = header_fill
    cell.border = thin_border
    current_row += 1

    for goal in short_term_goals:
        ws.merge_cells(f'A{current_row}:F{current_row}')
        cell = ws[f'A{current_row}']
        cell.value = f"• {goal.get('goal', '')}"
        cell.font = normal_font
        cell.alignment = left_align
        cell.border = thin_border
        ws.row_dimensions[current_row].height = 30
        current_row += 1

    current_row += 1

    # Support Items Table Header
    headers = ['項目', '支援目標\n（具体的な到達目標）', '支援内容\n（内容・支援の提供上のポイント）',
               '達成\n時期', '担当者・\n提供機関', '優先\n順位']

    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=current_row, column=col_idx)
        cell.value = header
        cell.font = Font(name='Meiryo UI', size=9, bold=True)
        cell.alignment = center_align
        cell.fill = header_fill
        cell.border = thin_border

    current_row += 1

    # Support Items (5 domains)
    if plan_data:
        support_items = get_field_value(
            plan_data,
            'support_items',
            assessment_v1.get('support_items', [])
        )
    else:
        support_items = assessment_v1.get('support_items', [])

    if not isinstance(support_items, list):
        support_items = []

    for item in support_items:
        start_row = current_row

        # Category
        ws[f'A{current_row}'] = item.get('category', '')
        ws[f'A{current_row}'].font = small_font
        ws[f'A{current_row}'].alignment = center_align
        ws[f'A{current_row}'].border = thin_border

        # Target
        ws[f'B{current_row}'] = item.get('target', '')
        ws[f'B{current_row}'].font = small_font
        ws[f'B{current_row}'].alignment = left_align
        ws[f'B{current_row}'].border = thin_border

        # Methods
        methods = item.get('methods', [])
        methods_text = '\n'.join([f"• {method}" for method in methods])
        ws[f'C{current_row}'] = methods_text
        ws[f'C{current_row}'].font = small_font
        ws[f'C{current_row}'].alignment = left_align
        ws[f'C{current_row}'].border = thin_border

        # Timeline
        ws[f'D{current_row}'] = item.get('timeline', '')
        ws[f'D{current_row}'].font = small_font
        ws[f'D{current_row}'].alignment = center_align
        ws[f'D{current_row}'].border = thin_border

        # Staff
        ws[f'E{current_row}'] = item.get('staff', '')
        ws[f'E{current_row}'].font = small_font
        ws[f'E{current_row}'].alignment = left_align
        ws[f'E{current_row}'].border = thin_border

        # Priority
        ws[f'F{current_row}'] = item.get('priority', '')
        ws[f'F{current_row}'].font = small_font
        ws[f'F{current_row}'].alignment = center_align
        ws[f'F{current_row}'].border = thin_border

        ws.row_dimensions[current_row].height = 60
        current_row += 1

    # Family Support
    if plan_data:
        family_support = get_field_value(
            plan_data,
            'family_support',
            assessment_v1.get('family_support', {})
        )
    else:
        family_support = assessment_v1.get('family_support', {})

    if not isinstance(family_support, dict):
        family_support = {}

    if family_support:
        ws[f'A{current_row}'] = '家族支援'
        ws[f'A{current_row}'].font = small_font
        ws[f'A{current_row}'].alignment = center_align
        ws[f'A{current_row}'].border = thin_border

        ws[f'B{current_row}'] = family_support.get('goal', '')
        ws[f'B{current_row}'].font = small_font
        ws[f'B{current_row}'].alignment = left_align
        ws[f'B{current_row}'].border = thin_border

        methods = family_support.get('methods', [])
        methods_text = '\n'.join([f"• {method}" for method in methods])
        ws[f'C{current_row}'] = methods_text
        ws[f'C{current_row}'].font = small_font
        ws[f'C{current_row}'].alignment = left_align
        ws[f'C{current_row}'].border = thin_border

        ws[f'D{current_row}'] = family_support.get('timeline', '')
        ws[f'D{current_row}'].font = small_font
        ws[f'D{current_row}'].alignment = center_align
        ws[f'D{current_row}'].border = thin_border

        ws[f'E{current_row}'] = '心理担当職員\n保護者'
        ws[f'E{current_row}'].font = small_font
        ws[f'E{current_row}'].alignment = left_align
        ws[f'E{current_row}'].border = thin_border

        ws[f'F{current_row}'] = ''
        ws[f'F{current_row}'].border = thin_border

        ws.row_dimensions[current_row].height = 50
        current_row += 1

    # Transition Support
    if plan_data:
        transition_support = get_field_value(
            plan_data,
            'transition_support',
            assessment_v1.get('transition_support', {})
        )
    else:
        transition_support = assessment_v1.get('transition_support', {})

    if not isinstance(transition_support, dict):
        transition_support = {}

    if transition_support:
        ws[f'A{current_row}'] = '移行支援'
        ws[f'A{current_row}'].font = small_font
        ws[f'A{current_row}'].alignment = center_align
        ws[f'A{current_row}'].border = thin_border

        ws[f'B{current_row}'] = transition_support.get('goal', '')
        ws[f'B{current_row}'].font = small_font
        ws[f'B{current_row}'].alignment = left_align
        ws[f'B{current_row}'].border = thin_border

        methods = transition_support.get('methods', [])
        methods_text = '\n'.join([f"• {method}" for method in methods])
        ws[f'C{current_row}'] = methods_text
        ws[f'C{current_row}'].font = small_font
        ws[f'C{current_row}'].alignment = left_align
        ws[f'C{current_row}'].border = thin_border

        ws[f'D{current_row}'] = transition_support.get('timeline', '')
        ws[f'D{current_row}'].font = small_font
        ws[f'D{current_row}'].alignment = center_align
        ws[f'D{current_row}'].border = thin_border

        partner = transition_support.get('partner_organization', '')
        ws[f'E{current_row}'] = f'児童発達支援管理責任者\n{partner}'
        ws[f'E{current_row}'].font = small_font
        ws[f'E{current_row}'].alignment = left_align
        ws[f'E{current_row}'].border = thin_border

        ws[f'F{current_row}'] = ''
        ws[f'F{current_row}'].border = thin_border

        ws.row_dimensions[current_row].height = 50
        current_row += 1

    # Footer
    current_row += 2
    ws.merge_cells(f'A{current_row}:C{current_row}')
    ws[f'A{current_row}'] = '提供する支援内容について、本計画書に基づき説明しました。'
    ws[f'A{current_row}'].font = small_font

    ws.merge_cells(f'D{current_row}:F{current_row}')
    ws[f'D{current_row}'] = '本計画書に基づき支援の説明を受け、内容に同意しました。'
    ws[f'D{current_row}'].font = small_font

    current_row += 1
    ws.merge_cells(f'A{current_row}:C{current_row}')
    ws[f'A{current_row}'] = '児童発達支援管理責任者氏名：'
    ws[f'A{current_row}'].font = small_font

    ws.merge_cells(f'D{current_row}:F{current_row}')
    ws[f'D{current_row}'] = f'{datetime.now().strftime("%Y年%m月%d日")}　（保護者署名）'
    ws[f'D{current_row}'].font = small_font


def generate_support_details_page2(
    ws,
    assessment_v1: dict,
    session_data: dict = None,
    plan_data: dict = None
):
    """
    Generate support details page (2/2) with 7-column table

    Args:
        ws: Worksheet object
        assessment_v1: Assessment data (fallback)
        session_data: Session data (child profile, etc.)
        plan_data: Optional business_support_plans record (2-column structure)
    """

    # Set column widths
    ws.column_dimensions['A'].width = 12   # 項目
    ws.column_dimensions['B'].width = 35   # 具体的な到達目標
    ws.column_dimensions['C'].width = 45   # 具体的な支援内容・5領域との関係性等
    ws.column_dimensions['D'].width = 12   # 達成時期
    ws.column_dimensions['E'].width = 25   # 提供期間
    ws.column_dimensions['F'].width = 30   # 留意事項
    ws.column_dimensions['G'].width = 10   # 優先順位

    # Styles
    header_font = Font(name='Meiryo UI', size=14, bold=True)
    normal_font = Font(name='Meiryo UI', size=10)
    small_font = Font(name='Meiryo UI', size=9)
    header_cell_font = Font(name='Meiryo UI', size=10, bold=True)

    center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    left_align = Alignment(horizontal='left', vertical='top', wrap_text=True)

    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    header_fill = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')

    current_row = 1

    # Title
    ws.merge_cells(f'A{current_row}:G{current_row}')
    cell = ws[f'A{current_row}']
    cell.value = '個別支援計画書（2/2ページ）'
    cell.font = header_font
    cell.alignment = center_align
    current_row += 1

    # Child name and date
    child_name = '〇〇 〇〇'
    if session_data:
        child_name = session_data.get('subject_name', child_name)

    if child_name == '〇〇 〇〇':
        child_profile = assessment_v1.get('child_profile', {})
        child_name = child_profile.get('name', child_name)

    ws.merge_cells(f'A{current_row}:D{current_row}')
    ws[f'A{current_row}'] = f'利用児氏名：{child_name}'
    ws[f'A{current_row}'].font = normal_font
    ws[f'A{current_row}'].alignment = left_align

    ws.merge_cells(f'E{current_row}:G{current_row}')
    ws[f'E{current_row}'] = f'作成日：{datetime.now().strftime("%Y年%m月%d日")}'
    ws[f'E{current_row}'].font = normal_font
    ws[f'E{current_row}'].alignment = Alignment(horizontal='right', vertical='center')
    current_row += 2

    # Table headers (7 columns)
    headers = [
        '項目',
        '具体的な到達目標',
        '具体的な支援内容・5領域との関係性等',
        '達成時期',
        '提供期間',
        '留意事項',
        '優先順位'
    ]

    for col_idx, header_text in enumerate(headers, start=1):
        cell = ws.cell(row=current_row, column=col_idx)
        cell.value = header_text
        cell.font = header_cell_font
        cell.alignment = center_align
        cell.fill = header_fill
        cell.border = thin_border

    ws.row_dimensions[current_row].height = 40
    current_row += 1

    # Data rows - Extract support items with 2-column priority logic
    if plan_data:
        support_items_raw = get_field_value(
            plan_data,
            'support_items',
            assessment_v1.get('support_items', [])
        )
    else:
        support_items_raw = assessment_v1.get('support_items', [])

    if not isinstance(support_items_raw, list):
        support_items_raw = []

    # Convert to 7-column table structure
    support_items = []
    for idx, item in enumerate(support_items_raw, start=1):
        # Extract methods array and join with line breaks
        methods = item.get('methods', [])
        if isinstance(methods, list):
            methods_text = '\n'.join([f"• {method}" for method in methods])
        else:
            methods_text = str(methods) if methods else ''

        support_items.append({
            '項目': item.get('category', '本人支援'),
            '具体的な到達目標': item.get('target', ''),
            '具体的な支援内容・5領域との関係性等': methods_text,
            '達成時期': item.get('timeline', '6ヶ月'),
            '提供期間': item.get('staff', '児童発達支援管理責任者、全職員'),
            '留意事項': item.get('notes', ''),
            '優先順位': str(item.get('priority', idx))
        })

    # If no support items, add empty rows for manual entry
    if not support_items:
        support_items = [
            {
                '項目': '',
                '具体的な到達目標': '',
                '具体的な支援内容・5領域との関係性等': '',
                '達成時期': '',
                '提供期間': '',
                '留意事項': '',
                '優先順位': ''
            }
            for _ in range(5)
        ]

    # Write data rows
    for item in support_items:
        row_height = max(60, len(item['具体的な支援内容・5領域との関係性等']) // 20 * 15)

        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=current_row, column=col_idx)
            cell.value = item.get(header, '')
            cell.font = normal_font
            cell.alignment = left_align if col_idx in [2, 3, 6] else center_align
            cell.border = thin_border

        ws.row_dimensions[current_row].height = row_height
        current_row += 1

    # Add empty rows for manual entry
    for _ in range(5):
        for col_idx in range(1, 8):
            cell = ws.cell(row=current_row, column=col_idx)
            cell.value = ''
            cell.border = thin_border
            cell.alignment = left_align

        ws.row_dimensions[current_row].height = 60
        current_row += 1


def generate_support_schedule(ws, assessment_v1: dict, session_data: dict = None):
    """Generate support schedule sheet (appendix)"""

    # Set column widths
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 18
    ws.column_dimensions['C'].width = 18
    ws.column_dimensions['D'].width = 18
    ws.column_dimensions['E'].width = 18
    ws.column_dimensions['F'].width = 18
    ws.column_dimensions['G'].width = 18
    ws.column_dimensions['H'].width = 18

    # Styles
    header_font = Font(name='Meiryo UI', size=14, bold=True)
    normal_font = Font(name='Meiryo UI', size=10)
    small_font = Font(name='Meiryo UI', size=9)

    center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    left_align = Alignment(horizontal='left', vertical='top', wrap_text=True)

    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    header_fill = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')

    current_row = 1

    # Title
    ws.merge_cells(f'A{current_row}:H{current_row}')
    cell = ws[f'A{current_row}']
    cell.value = '個別支援計画別表'
    cell.font = header_font
    cell.alignment = center_align
    current_row += 2

    # Child name and date - prioritize subjects table data
    child_name = '〇〇 〇〇'
    if session_data:
        child_name = session_data.get('subject_name', child_name)

    # Fallback to assessment_v1 child_profile
    if child_name == '〇〇 〇〇':
        child_profile = assessment_v1.get('child_profile', {})
        child_name = child_profile.get('name', child_name)

    ws.merge_cells(f'A{current_row}:E{current_row}')
    ws[f'A{current_row}'] = f'利用児氏名：{child_name}'
    ws[f'A{current_row}'].font = normal_font
    ws[f'A{current_row}'].fill = header_fill
    ws[f'A{current_row}'].border = thin_border

    ws.merge_cells(f'F{current_row}:H{current_row}')
    ws[f'F{current_row}'] = f'作成日：{datetime.now().strftime("%Y年%m月%d日")}'
    ws[f'F{current_row}'].font = normal_font
    ws[f'F{current_row}'].alignment = Alignment(horizontal='right', vertical='center')
    current_row += 2

    # Table headers
    days = ['月', '火', '水', '木', '金', '土', '日・祝日']
    ws[f'A{current_row}'] = ''
    ws[f'A{current_row}'].fill = header_fill
    ws[f'A{current_row}'].border = thin_border

    for col_idx, day in enumerate(days, start=2):
        cell = ws.cell(row=current_row, column=col_idx)
        cell.value = day
        cell.font = Font(name='Meiryo UI', size=10, bold=True)
        cell.alignment = center_align
        cell.fill = header_fill
        cell.border = thin_border

    current_row += 1

    # Rows (dummy data)
    row_labels = [
        ('提供時間', [
            '利用開始・終了時間',
            '',
        ]),
        ('延長支援時間\n※延長支援時間は、\n支援前・支援後\nそれぞれ1時間以上から', [
            '【支援前】延長支援時間',
            '【支援後】延長支援時間',
            '',
        ]),
    ]

    for label, sub_labels in row_labels:
        start_row = current_row
        for sub_label in sub_labels:
            ws[f'A{current_row}'] = label if current_row == start_row else ''
            ws[f'A{current_row}'].font = small_font
            ws[f'A{current_row}'].alignment = left_align
            ws[f'A{current_row}'].fill = header_fill
            ws[f'A{current_row}'].border = thin_border

            # Dummy data for each day
            for col_idx in range(2, 9):
                cell = ws.cell(row=current_row, column=col_idx)
                if sub_label == '利用開始・終了時間':
                    cell.value = '10時00分～15時00分\n5時00分' if col_idx in [2, 4, 6] else '～\n0:00'
                elif '【支援前】' in sub_label:
                    cell.value = '9時00分～10時00分' if col_idx in [2, 4, 6] else '～'
                elif '【支援後】' in sub_label:
                    cell.value = '15時00分～16時00分\n2時00分' if col_idx in [2, 4, 6] else '～'
                else:
                    cell.value = ''

                cell.font = small_font
                cell.alignment = center_align
                cell.border = thin_border

            ws.row_dimensions[current_row].height = 30
            current_row += 1

        if start_row < current_row - 1:
            ws.merge_cells(f'A{start_row}:A{current_row - 1}')

    # Extended support reason
    ws.merge_cells(f'A{current_row}:H{current_row}')
    ws[f'A{current_row}'] = '延長を必要とする理由'
    ws[f'A{current_row}'].font = Font(name='Meiryo UI', size=10, bold=True)
    ws[f'A{current_row}'].fill = header_fill
    ws[f'A{current_row}'].border = thin_border
    current_row += 1

    ws.merge_cells(f'A{current_row}:H{current_row + 1}')
    ws[f'A{current_row}'] = '例①）月・水・金については、保護者の就労を理由に支援前・支援後それぞれ1時間ずつの延長支援を行う。\n例②）保護者の職場の繁忙期（3月）については、月・水・金の支援後の延長支援時間が2時間になる日も生じることが想定されるため、保護者と連携を図りながら必要に応じて延長支援を行う。'
    ws[f'A{current_row}'].font = small_font
    ws[f'A{current_row}'].alignment = left_align
    ws[f'A{current_row}'].border = thin_border
    ws.row_dimensions[current_row].height = 50
    current_row += 2

    # Special notes
    ws.merge_cells(f'A{current_row}:H{current_row}')
    ws[f'A{current_row}'] = '特記事項'
    ws[f'A{current_row}'].font = Font(name='Meiryo UI', size=10, bold=True)
    ws[f'A{current_row}'].fill = header_fill
    ws[f'A{current_row}'].border = thin_border
    current_row += 1

    ws.merge_cells(f'A{current_row}:H{current_row}')
    ws[f'A{current_row}'] = ''
    ws[f'A{current_row}'].border = thin_border
    ws.row_dimensions[current_row].height = 40


def extract_assessment_v1(assessment_result: dict) -> dict:
    """
    Extract assessment_v1 from various data formats

    Handles:
    - Direct: {"assessment_v1": {...}}
    - Wrapped: {"summary": "```json\n{...}\n```"}
    """
    if not assessment_result:
        return {}

    # Case 1: Direct structure
    if 'assessment_v1' in assessment_result:
        return assessment_result['assessment_v1']

    # Case 2: Wrapped in summary
    if 'summary' in assessment_result:
        import re
        import json

        summary_text = assessment_result['summary']
        # Extract JSON from markdown code block
        json_match = re.search(r'```json\s*\n([\s\S]*?)\n```', summary_text)
        if json_match:
            try:
                parsed = json.loads(json_match.group(1))
                return parsed.get('assessment_v1', {})
            except json.JSONDecodeError:
                pass

    return {}


def get_field_value(plan: dict, field_prefix: str, fallback=None):
    """
    Get field value with 2-column priority logic

    Priority:
    1. {field_prefix}_user_edited (if not null)
    2. {field_prefix}_ai_generated
    3. fallback (if provided)

    Args:
        plan: business_support_plans record
        field_prefix: Field name without suffix (e.g., "child_intention")
        fallback: Optional fallback value

    Returns:
        Field value (str, list, dict, or None)
    """
    if not plan:
        return fallback

    user_edited = plan.get(f'{field_prefix}_user_edited')
    ai_generated = plan.get(f'{field_prefix}_ai_generated')

    if user_edited is not None:
        return user_edited
    if ai_generated is not None:
        return ai_generated
    return fallback


def generate_support_plan_excel_from_plan(
    plan_data: dict,
    subject_data: dict = None
) -> BytesIO:
    """
    Generate Individual Support Plan Excel from plan_data only (session not required)

    Args:
        plan_data: business_support_plans record (2-column structure)
        subject_data: Optional subject info for child name/age

    Returns:
        BytesIO: Excel file binary data
    """
    wb = Workbook()

    # Sheet 1: Main Support Plan
    ws1 = wb.active
    ws1.title = "個別支援計画書1"
    generate_main_support_plan_from_plan(ws1, plan_data, subject_data)

    # Sheet 2: Support Details (Page 2)
    ws2 = wb.create_sheet(title="個別支援計画書2")
    generate_support_details_page2_from_plan(ws2, plan_data, subject_data)

    # Sheet 3: Support Schedule (Appendix)
    ws3 = wb.create_sheet(title="支援計画別表")
    generate_support_schedule_from_plan(ws3, plan_data, subject_data)

    # Save to BytesIO
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    return output


def generate_main_support_plan_from_plan(
    ws,
    plan_data: dict,
    subject_data: dict = None
):
    """
    Generate main support plan sheet from plan_data only
    Layout matches the UI exactly (4-column grid)
    """
    # 4-column layout: A=label, B=value, C=label, D=value
    ws.column_dimensions['A'].width = 18
    ws.column_dimensions['B'].width = 40
    ws.column_dimensions['C'].width = 18
    ws.column_dimensions['D'].width = 30

    # Styles
    title_font = Font(name='Meiryo UI', size=16, bold=True)
    normal_font = Font(name='Meiryo UI', size=10)
    label_font = Font(name='Meiryo UI', size=10, bold=True)
    sublabel_font = Font(name='Meiryo UI', size=9, bold=True)
    name_font = Font(name='Meiryo UI', size=13, bold=True)

    center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    left_align = Alignment(horizontal='left', vertical='center', wrap_text=True)
    left_top_align = Alignment(horizontal='left', vertical='top', wrap_text=True)

    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    label_fill = PatternFill(start_color='F9F9F9', end_color='F9F9F9', fill_type='solid')

    def write_label(row, col, text):
        cell = ws.cell(row=row, column=col)
        cell.value = text
        cell.font = label_font
        cell.alignment = left_align
        cell.fill = label_fill
        cell.border = thin_border

    def write_value(row, col, text, font=None):
        cell = ws.cell(row=row, column=col)
        cell.value = text or ''
        cell.font = font or normal_font
        cell.alignment = left_align
        cell.border = thin_border

    # Extract data
    child_name = subject_data.get('name', '') if subject_data else ''
    child_age = subject_data.get('age', '') if subject_data else ''
    birth_date = subject_data.get('birth_date', '') if subject_data else ''

    current_row = 1

    # === Title ===
    ws.merge_cells(f'A{current_row}:D{current_row}')
    cell = ws[f'A{current_row}']
    cell.value = '個別支援計画書'
    cell.font = title_font
    cell.alignment = center_align
    ws.row_dimensions[current_row].height = 40
    current_row += 2

    # === Row 1: facility_name | birth_date ===
    write_label(current_row, 1, '事業所名')
    facility_name = get_field_value(plan_data, 'facility_name', '') or plan_data.get('facility_name', '') or 'ヨリドコロ横浜白楽'
    write_value(current_row, 2, facility_name)
    write_label(current_row, 3, '生年月日')
    age_text = f' ({child_age}歳)' if child_age else ''
    write_value(current_row, 4, f'{birth_date}{age_text}' if birth_date else '')
    ws.row_dimensions[current_row].height = 30
    current_row += 1

    # === Row 2: manager_name | created_at ===
    write_label(current_row, 1, '計画作成者')
    manager_name = plan_data.get('manager_name', '') or '児童発達支援管理責任者 山田太郎'
    write_value(current_row, 2, manager_name)
    write_label(current_row, 3, '計画作成日')
    created_at = plan_data.get('created_at', '')
    if created_at:
        try:
            dt = datetime.fromisoformat(created_at.replace('Z', '+00:00'))
            created_at = dt.strftime('%Y/%m/%d')
        except (ValueError, TypeError):
            pass
    write_value(current_row, 4, created_at)
    ws.row_dimensions[current_row].height = 30
    current_row += 1

    # === Row 3: child_name | monitoring_period ===
    write_label(current_row, 1, '利用者氏名')
    write_value(current_row, 2, f'{child_name} 様' if child_name else '', font=name_font)
    write_label(current_row, 3, 'ﾓﾆﾀﾘﾝｸﾞ期間')
    m_start = plan_data.get('monitoring_start', '')
    m_end = plan_data.get('monitoring_end', '')
    monitoring_text = f'{m_start} 〜 {m_end}' if m_start or m_end else ''
    write_value(current_row, 4, monitoring_text)
    ws.row_dimensions[current_row].height = 35
    current_row += 1

    # === Intentions header (full width) ===
    ws.merge_cells(f'A{current_row}:D{current_row}')
    cell = ws[f'A{current_row}']
    cell.value = '利用者及びその家族の生活に対する意向・ニーズ（生活全般の質を向上させるための課題）'
    cell.font = label_font
    cell.fill = label_fill
    cell.alignment = left_align
    cell.border = thin_border
    ws.row_dimensions[current_row].height = 30
    current_row += 1

    # === Intention: child ===
    child_intention = get_field_value(plan_data, 'child_intention', '')
    write_label(current_row, 1, 'ご本人')
    ws[f'A{current_row}'].font = sublabel_font
    ws.merge_cells(f'B{current_row}:D{current_row}')
    write_value(current_row, 2, child_intention)
    ws[f'B{current_row}'].alignment = left_top_align
    ws.row_dimensions[current_row].height = 40
    current_row += 1

    # === Intention: family ===
    family_intention = get_field_value(plan_data, 'family_intention', '')
    write_label(current_row, 1, 'ご家族')
    ws[f'A{current_row}'].font = sublabel_font
    ws.merge_cells(f'B{current_row}:D{current_row}')
    write_value(current_row, 2, family_intention)
    ws[f'B{current_row}'].alignment = left_top_align
    ws.row_dimensions[current_row].height = 40
    current_row += 1

    # === Service schedule (full width label + value) ===
    write_label(current_row, 1, '支援の標準的な\n提供時間等')
    ws[f'A{current_row}'].alignment = left_top_align
    ws.merge_cells(f'B{current_row}:D{current_row}')
    write_value(current_row, 2, plan_data.get('service_schedule', ''))
    ws[f'B{current_row}'].alignment = left_top_align
    ws.row_dimensions[current_row].height = 40
    current_row += 1

    # === Notes (full width label + value) ===
    write_label(current_row, 1, '留意点・備考')
    ws.merge_cells(f'B{current_row}:D{current_row}')
    write_value(current_row, 2, plan_data.get('notes', ''))
    ws[f'B{current_row}'].alignment = left_top_align
    ws.row_dimensions[current_row].height = 40
    current_row += 1

    # === General policy (full width label + value) ===
    general_policy = get_field_value(plan_data, 'general_policy', '')
    write_label(current_row, 1, '総合的な\n支援の方針')
    ws[f'A{current_row}'].alignment = left_top_align
    ws.merge_cells(f'B{current_row}:D{current_row}')
    write_value(current_row, 2, general_policy)
    ws[f'B{current_row}'].alignment = left_top_align
    ws.row_dimensions[current_row].height = 80
    current_row += 1

    # === Long-term goal | period ===
    lt_goal = get_field_value(plan_data, 'long_term_goal', '')
    lt_period = get_field_value(plan_data, 'long_term_period', '1年')
    write_label(current_row, 1, '長期目標')
    write_value(current_row, 2, lt_goal)
    ws[f'B{current_row}'].alignment = left_top_align
    write_label(current_row, 3, '期間')
    write_value(current_row, 4, lt_period)
    ws.row_dimensions[current_row].height = 50
    current_row += 1

    # === Short-term goal | period ===
    short_term_goals = get_field_value(plan_data, 'short_term_goals', [])
    st_goal_text = ''
    st_period = '6ヶ月'
    if isinstance(short_term_goals, list) and len(short_term_goals) > 0:
        first = short_term_goals[0]
        st_goal_text = first.get('goal', '') if isinstance(first, dict) else str(first)
        st_period = first.get('timeline', '6ヶ月') if isinstance(first, dict) else '6ヶ月'
    # Also check single short_term_goal field
    st_goal_override = plan_data.get('short_term_goal')
    if st_goal_override:
        st_goal_text = st_goal_override
    st_period_override = plan_data.get('short_term_period')
    if st_period_override:
        st_period = st_period_override

    write_label(current_row, 1, '短期目標')
    write_value(current_row, 2, st_goal_text)
    ws[f'B{current_row}'].alignment = left_top_align
    write_label(current_row, 3, '期間')
    write_value(current_row, 4, st_period)
    ws.row_dimensions[current_row].height = 50
    current_row += 1


def generate_support_details_page2_from_plan(
    ws,
    plan_data: dict,
    subject_data: dict = None
):
    """
    Generate support details page (2/2) from plan_data only
    """
    # Set column widths
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 35
    ws.column_dimensions['C'].width = 45
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 25
    ws.column_dimensions['F'].width = 30
    ws.column_dimensions['G'].width = 10

    # Styles
    header_font = Font(name='Meiryo UI', size=14, bold=True)
    normal_font = Font(name='Meiryo UI', size=10)
    header_cell_font = Font(name='Meiryo UI', size=10, bold=True)

    center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    left_align = Alignment(horizontal='left', vertical='top', wrap_text=True)

    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    header_fill = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')

    current_row = 1

    # Title
    ws.merge_cells(f'A{current_row}:G{current_row}')
    cell = ws[f'A{current_row}']
    cell.value = '個別支援計画書（2/2ページ）'
    cell.font = header_font
    cell.alignment = center_align
    current_row += 1

    # Child name
    child_name = subject_data.get('name', '〇〇 〇〇') if subject_data else '〇〇 〇〇'

    ws.merge_cells(f'A{current_row}:D{current_row}')
    ws[f'A{current_row}'] = f'利用児氏名：{child_name}'
    ws[f'A{current_row}'].font = normal_font
    ws[f'A{current_row}'].alignment = left_align

    ws.merge_cells(f'E{current_row}:G{current_row}')
    ws[f'E{current_row}'] = f'作成日：{datetime.now().strftime("%Y年%m月%d日")}'
    ws[f'E{current_row}'].font = normal_font
    ws[f'E{current_row}'].alignment = Alignment(horizontal='right', vertical='center')
    current_row += 2

    # Table headers
    headers = ['項目', '具体的な到達目標', '具体的な支援内容・5領域との関係性等',
               '達成時期', '提供期間', '留意事項', '優先順位']

    for col_idx, header_text in enumerate(headers, start=1):
        cell = ws.cell(row=current_row, column=col_idx)
        cell.value = header_text
        cell.font = header_cell_font
        cell.alignment = center_align
        cell.fill = header_fill
        cell.border = thin_border

    ws.row_dimensions[current_row].height = 40
    current_row += 1

    # Data rows
    support_items = get_field_value(plan_data, 'support_items', [])
    if not isinstance(support_items, list):
        support_items = []

    for idx, item in enumerate(support_items, start=1):
        methods = item.get('methods', [])
        if isinstance(methods, list):
            methods_text = '\n'.join([f"• {method}" for method in methods])
        else:
            methods_text = str(methods) if methods else ''

        row_data = {
            '項目': item.get('category', '本人支援'),
            '具体的な到達目標': item.get('target', ''),
            '具体的な支援内容・5領域との関係性等': methods_text,
            '達成時期': item.get('timeline', '6ヶ月'),
            '提供期間': item.get('staff', '児童発達支援管理責任者、全職員'),
            '留意事項': item.get('notes', ''),
            '優先順位': str(item.get('priority', idx))
        }

        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=current_row, column=col_idx)
            cell.value = row_data.get(header, '')
            cell.font = normal_font
            cell.alignment = left_align if col_idx in [2, 3, 6] else center_align
            cell.border = thin_border

        ws.row_dimensions[current_row].height = 60
        current_row += 1

    # Add empty rows for manual entry
    for _ in range(5):
        for col_idx in range(1, 8):
            cell = ws.cell(row=current_row, column=col_idx)
            cell.value = ''
            cell.border = thin_border
            cell.alignment = left_align
        ws.row_dimensions[current_row].height = 60
        current_row += 1

    # Footer (agreement / consent)
    current_row += 2
    ws.merge_cells(f'A{current_row}:G{current_row}')
    ws[f'A{current_row}'] = '提供する支援内容について、本計画書に基づき説明を受け、内容に同意しました。'
    ws[f'A{current_row}'].font = normal_font
    current_row += 2

    ws.merge_cells(f'A{current_row}:C{current_row}')
    ws[f'A{current_row}'] = '説明者：'
    ws[f'A{current_row}'].font = normal_font

    ws.merge_cells(f'E{current_row}:G{current_row}')
    ws[f'E{current_row}'] = f'説明・同意日：{datetime.now().strftime("%Y年%m月%d日")}'
    ws[f'E{current_row}'].font = normal_font
    ws[f'E{current_row}'].alignment = Alignment(horizontal='right', vertical='center')
    current_row += 1

    ws.merge_cells(f'E{current_row}:G{current_row}')
    ws[f'E{current_row}'] = '保護者氏名：　　　　　　　（自署または捺印）'
    ws[f'E{current_row}'].font = normal_font
    ws[f'E{current_row}'].alignment = Alignment(horizontal='right', vertical='center')


def generate_support_schedule_from_plan(
    ws,
    plan_data: dict,
    subject_data: dict = None
):
    """
    Generate support schedule sheet (appendix) from plan_data only
    """
    # Set column widths
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 18
    ws.column_dimensions['C'].width = 18
    ws.column_dimensions['D'].width = 18
    ws.column_dimensions['E'].width = 18
    ws.column_dimensions['F'].width = 18
    ws.column_dimensions['G'].width = 18
    ws.column_dimensions['H'].width = 18

    # Styles
    header_font = Font(name='Meiryo UI', size=14, bold=True)
    normal_font = Font(name='Meiryo UI', size=10)
    small_font = Font(name='Meiryo UI', size=9)

    center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    left_align = Alignment(horizontal='left', vertical='top', wrap_text=True)

    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    header_fill = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')

    current_row = 1

    # Title
    ws.merge_cells(f'A{current_row}:H{current_row}')
    cell = ws[f'A{current_row}']
    cell.value = '個別支援計画別表'
    cell.font = header_font
    cell.alignment = center_align
    current_row += 2

    # Child name
    child_name = subject_data.get('name', '〇〇 〇〇') if subject_data else '〇〇 〇〇'

    ws.merge_cells(f'A{current_row}:E{current_row}')
    ws[f'A{current_row}'] = f'利用児氏名：{child_name}'
    ws[f'A{current_row}'].font = normal_font
    ws[f'A{current_row}'].fill = header_fill
    ws[f'A{current_row}'].border = thin_border

    ws.merge_cells(f'F{current_row}:H{current_row}')
    ws[f'F{current_row}'] = f'作成日：{datetime.now().strftime("%Y年%m月%d日")}'
    ws[f'F{current_row}'].font = normal_font
    ws[f'F{current_row}'].alignment = Alignment(horizontal='right', vertical='center')
    current_row += 2

    # Table headers
    days = ['月', '火', '水', '木', '金', '土', '日・祝日']
    ws[f'A{current_row}'] = ''
    ws[f'A{current_row}'].fill = header_fill
    ws[f'A{current_row}'].border = thin_border

    for col_idx, day in enumerate(days, start=2):
        cell = ws.cell(row=current_row, column=col_idx)
        cell.value = day
        cell.font = Font(name='Meiryo UI', size=10, bold=True)
        cell.alignment = center_align
        cell.fill = header_fill
        cell.border = thin_border

    current_row += 1

    # Empty rows for schedule
    row_labels = ['提供時間', '延長支援時間']
    for label in row_labels:
        ws[f'A{current_row}'] = label
        ws[f'A{current_row}'].font = small_font
        ws[f'A{current_row}'].alignment = left_align
        ws[f'A{current_row}'].fill = header_fill
        ws[f'A{current_row}'].border = thin_border

        for col_idx in range(2, 9):
            cell = ws.cell(row=current_row, column=col_idx)
            cell.value = ''
            cell.font = small_font
            cell.alignment = center_align
            cell.border = thin_border

        ws.row_dimensions[current_row].height = 40
        current_row += 1

    # Special notes
    current_row += 1
    ws.merge_cells(f'A{current_row}:H{current_row}')
    ws[f'A{current_row}'] = '特記事項'
    ws[f'A{current_row}'].font = Font(name='Meiryo UI', size=10, bold=True)
    ws[f'A{current_row}'].fill = header_fill
    ws[f'A{current_row}'].border = thin_border
    current_row += 1

    ws.merge_cells(f'A{current_row}:H{current_row}')
    ws[f'A{current_row}'] = ''
    ws[f'A{current_row}'].border = thin_border
    ws.row_dimensions[current_row].height = 60
